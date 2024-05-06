import os
import sqlite3
import openpyxl
import csv
import pandas as pd
from datetime import datetime
from datetime import timedelta


# variables de tipo fecha
fecha_actual = datetime.today().date()
fecha_60_dias = timedelta(days=60)
fecha_limite = fecha_actual + fecha_60_dias


# def funcionando
def limpiar_consola():
    pass
    # os.system("cls")


# def funcionando
def creacion_tablas():
    # Creeando las tablas
    try:
        with sqlite3.connect("base_datos_ev3.db") as conn:
            cursor = conn.cursor()
            cursor.execute(
                "CREATE TABLE IF NOT EXISTS registro_pacientes \
                (id_paciente     INTEGER PRIMARY KEY,\
                primer_apellido     TEXT NOT NULL, \
                segundo_apellido    TEXT NULL, \
                nombre              TEXT NOT NULL, \
                fecha_nacimiento    TIMESTAMP NOT NULL, \
                sexo                TEXT NOT NULL);"
            )
            # print("Tabla 'registro_pacientes' creada correctamente.")

            cursor.execute(
                "CREATE TABLE IF NOT EXISTS citas \
                (id_folio INTEGER PRIMARY KEY, \
                id_paciente INTEGER NOT NULL, \
                fecha_cita TIMESTAMP NOT NULL, \
                turno TEXT NOT NULL, \
                hora_llegada TIMESTAMP NULL, \
                peso REAL NULL, \
                estatura REAL NULL, \
                presion_arterial TEXT NULL, \
                diagnostico TEXT NULL, \
                edad INTEGER NULL, \
                FOREIGN KEY (id_paciente) REFERENCES registro_pacientes(id_paciente));"
            )
            # print("Tabla 'citas' creada correctamente.")
    except sqlite3.Error as e:
        print(e)
    finally:
        if conn:
            conn.close()


# def funcionando
def registrar_pacientes():
    global fecha_actual
    while True:
        print("\n\n========REGISTRANDO AL PACIENTE=======")
        print("AVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")
        while True:
            primer_apellido = input(
                "\nIngresa el primer apellido del paciente: "
            ).upper()
            if primer_apellido == "*":
                break
            if primer_apellido == "":
                print("No puedes OMITIR este campo. Intenta de nuevo.")
                continue
            elif " " in primer_apellido:
                print(
                    "El apellido no puede contener espacios en blanco. Intente de nuevo."
                )
                continue
            elif not primer_apellido.isalpha():
                print("Debes ingresar valores de TEXTO. Intenta de nuevo.")
                continue
            else:
                break
        if primer_apellido == "*":
            break

        while True:
            segundo_apellido = input(
                "\nIngresa el segundo apellido del paciente: "
            ).upper()
            if segundo_apellido == "*":
                break
            # este campo si se puede omitir deacuerdo si el paciente no tiene segundo apellido
            if segundo_apellido == "":
                print("Apellido OMITIDO")
                segundo_apellido = "N/A"
                break
            elif " " in segundo_apellido:
                print(
                    "El apellido no puede contener espacios en blanco. Intente de nuevo."
                )
            elif not segundo_apellido.isalpha():
                print("Debes ingresar valores de TEXTO. Intenta de nuevo.")
                continue
            else:
                break
        if segundo_apellido == "*":
            break

        while True:
            nombre = input("\nIngrese el nombre del paciente: ").upper()
            if nombre == "*":
                break
            if nombre == "":
                print("No puede OMITIR este campo. Intente de vuevo.")
                continue
            elif not nombre.replace(" ", "").isalpha():
                print("Debes ingresar valores de TEXTO. Intenta de nuevo.")
                continue
            else:
                nombre = nombre.strip()
                break
        if nombre == "*":
            break

        while True:
            fecha_nacimiento = input(
                "\nIngrese la fecha de nacimiento del paciente (MM/DD/YYYY): "
            )
            if fecha_nacimiento == "*":
                break
            if fecha_nacimiento == "":
                print("No puedes OMITIR este valor. Intente de nuevo.")
                continue
            try:
                fecha_nacimiento = datetime.strptime(
                    fecha_nacimiento, "%m/%d/%Y"
                ).date()
                fecha_actual = datetime.today().date()
                if fecha_nacimiento <= fecha_actual:
                    # si entra en esta condicion la fecha es valida
                    break
                else:
                    print(
                        "La fecha de nacimiento debe ser menor a la fecha actual. Intente de nuevo."
                    )
                    continue
            except Exception:
                print(
                    "Debes de ingresar el formato correcto (MM/DD/YYYY). Intente de nuevo."
                )
        if fecha_nacimiento == "*":
            break

        while True:
            sexo = input(
                "\nIngresa el sexo del paciente (1. Hombre 2.Mujer 3. No contestar): "
            )
            if sexo == "*":
                break
            if sexo == "":
                print("No puedes omitir este campo")
                continue
            try:
                sexo = int(sexo)
            except ValueError:
                print("Ingresa valores Numericos")

            if sexo in [1, 2, 3]:
                if sexo == 1:
                    sexo = "H"
                elif sexo == 2:
                    sexo = "M"
                elif sexo == 3:
                    sexo = "N"
                break
            else:
                print("Debes seleccionar una opcion entre 1-3")
                continue
        if sexo == "*":
            break

        registro_pacientes = (
            primer_apellido,
            segundo_apellido,
            nombre,
            fecha_nacimiento,
            sexo,
        )
        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO registro_pacientes (primer_apellido, segundo_apellido, nombre , fecha_nacimiento, sexo)\
                    VALUES (?, ?, ?, ?, ?)",
                    registro_pacientes,
                )
                print("Datos insertados correctamente en el 'registro_pacientes'")
        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()

        break


# def en procesp
def programar_citas():
    global fecha_actual
    global fecha_60_dias
    global fecha_limite
    # conectando a la tabla de citas
    while True:
        # limpiar_consola()
        print("\nAVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")
        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id_paciente FROM registro_pacientes")
                resultado = cursor.fetchall()
        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()

        # verificando que contenga ids la sentencia, quiere decir si existen pacientes registrados
        if resultado:
            print("\n\n====CLAVES DE LOS PACIENTES REGISTRADOS===")

            # list compresion que contiene los id existentes en la tabla de registro_pacientes
            list_comp_ids = [id[0] for id in resultado]

            # iterando los valores que existen en la list comprehsion
            for id in list_comp_ids:
                print(f"{id:^40}")

            # bucle recibiendo id del paciente a programar
            while True:
                id_paciente_programar = input("\nIntroduza el id del paciente: ")
                if id_paciente_programar == "*":
                    break
                if id_paciente_programar == "":
                    print("No puedes OMITIR este campo. Intenta de nuevo.")
                    continue
                try:
                    id_paciente_programar = int(id_paciente_programar)
                    break
                except Exception:
                    print("Porfavor introduce valores NUMERICOS. Intenta de nuevo.")
                    continue

            if id_paciente_programar == "*":
                break

            if id_paciente_programar in list_comp_ids:
                # print("El id del paciente seleccionado se encuentra registrado")
                # bucle para la validacion de la fecha
                while True:
                    print(
                        f"\nAVISO: \n\t-La fecha mas distante que podra programar la cita es el dia {fecha_limite.strftime('%m/%d/%Y')} \
                        \n\t-No puede ser un dia Domingo"
                    )

                    fecha_cita = input("\nIntroduza la fecha de la cita (MM/DD/YYYY): ")
                    if fecha_cita == "*":
                        break
                    if fecha_cita == "":
                        print("No puedes OMITIR este campo. Intenta de nuevo.")
                        continue
                    if fecha_cita.isalpha():
                        print(
                            "Estas intentando introducir valores de TEXTO. Intenta de nuevo."
                        )
                        continue
                    if fecha_cita.isdigit():
                        print(
                            "Estas intentando introducir valores ENTEROS. Intenta de nuevo."
                        )
                        continue
                    try:
                        fecha_cita = datetime.strptime(fecha_cita, "%m/%d/%Y").date()
                    except Exception:
                        print(
                            "Porfavor introduce el formato valido (MM/DD/YYYY). Intenta de nuevo."
                        )
                        continue

                    if fecha_cita < fecha_actual:
                        print(
                            "La fecha de la cita debe ser mayor a la fecha actual. Intenta de nuevo."
                        )
                        continue
                    elif fecha_cita > fecha_limite:
                        print(
                            "La fecha de la cita tiene que estar dentro de los 60 dias a partir del dia de hoy. Intente de nuevo."
                        )
                        continue

                    validacion_fecha_cita_domingo = fecha_cita.weekday()
                    if validacion_fecha_cita_domingo == 6:
                        print(
                            f"La fecha proporcionada {fecha_cita} cae en un dia domingo."
                        )

                        decision_sabado = input(
                            "\nDeseas que sea el sábado inmediato antes de la fecha deseada (SI/NO): "
                        )
                        if decision_sabado == "*":
                            break
                        if decision_sabado.upper() == "SI":
                            fecha_cita = fecha_cita - timedelta(days=1)
                            print(
                                f"Su fecha quedo programada para el dia sabado {fecha_cita.strftime('%m/%d/%Y')}"
                            )
                        else:
                            print("No puedes progrmar una cita para el dia 'domingo'")
                            continue

                        if decision_sabado == "*":
                            break

                    #  bucle para validar el turno de la cita
                    while True:
                        turno = input(
                            "\nIngrese el turno de la cita puede ser de (1. Mañana 2. Mediodia 3. Tarde): "
                        )
                        if turno == "*":
                            break
                        if turno == "":
                            print("No puedes OMITIR este campo. Intenta de nuevo.")
                            continue
                        if turno.isalpha():
                            print(
                                "Estas intentando ingresar valores de TEXTO. Intenta de  nuevo."
                            )
                            continue
                        try:
                            turno = int(turno)
                        except Exception:
                            print(
                                "Porfavor introduce valores NUMERICOS. Intenta de nuevo."
                            )
                            continue
                        if turno in [1, 2, 3]:
                            if turno == 1:
                                turno = "MAÑANA"
                            elif turno == 2:
                                turno = "MEDIODIA"
                            elif turno == 3:
                                turno = "TARDE"
                            # termina el bucle de la validacion del turno
                            break
                        else:
                            print("Selecciona una opcion entre 1-3. Intente de nuevo")
                            continue

                    if turno == "*":
                        break

                    programacion_citas = (
                        id_paciente_programar,
                        fecha_cita,
                        turno,
                    )

                    try:
                        with sqlite3.connect("base_datos_ev3.db") as conn:
                            cursor = conn.cursor()
                            cursor.execute(
                                "INSERT INTO citas (id_paciente, fecha_cita, turno)\
                                VALUES (?, ?, ?)",
                                programacion_citas,
                            )
                            break
                    except sqlite3.Error as e:
                        print(e)
                    finally:
                        if conn:
                            conn.close()

            else:
                print(
                    f"El paciente con el id {id_paciente_programar} no se encuentra registrado"
                )

        else:
            print(
                "\nNo existen pacientes registrados no puedes programar una cita. Regresando al menu principal..."
            )
        # rompe el bucle principal que muestra las opciones del menu
        break


# def lista
def realizar_cita():
    while True:
        print("AVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")
        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id_folio FROM citas")
                resultado = cursor.fetchall()
        except sqlite3.Error as e:
            print(e)

        if resultado:
            # list compresion que obtiene los folios
            list_comp_folios = [folio[0] for folio in resultado]

            print("====FOLIOS DISPONIBLES====")
            for folio in list_comp_folios:
                print(f"{folio:^25}")

            # bucle para revisasr el folio
            while True:
                id_folio_realizar_cita = input(
                    "\nIngresa el folio para buscar la cita: "
                )
                if id_folio_realizar_cita == "*":
                    break
                if id_folio_realizar_cita == "":
                    print("No puedes OMITIR este campo. Intenta de nuevo.")
                    continue
                try:
                    id_folio_realizar_cita = int(id_folio_realizar_cita)
                    break
                except Exception:
                    print("Porfavor introduce valores NUMERICOS. Intenta de nuevo.")
                    continue

            if id_folio_realizar_cita == "*":
                break

            if id_folio_realizar_cita in list_comp_folios:
                # bucle para revisar el peso
                while True:
                    peso = input("\nIngrese el peso del paciente en kilogramos: ")
                    if peso == "*":
                        break
                    if peso == "":
                        print("No puedes OMITIR este valor. Intenta de nuevo.")
                        continue
                    if peso.isalpha():
                        print(
                            "Estas intentando introducir valores de TEXTO. Intenta de nuevo."
                        )
                        continue
                    try:
                        peso = float(peso)
                    except Exception:
                        print("Porfavor introduce valores NUMERICOS. Intenta de nuevo.")
                        continue
                    if peso > 0:
                        break
                    else:
                        print("No puedes ingresar un peso NEGATIVO. Intenta de nuevo.")
                        continue

                        # rompe el bucle para revisar el peso

                if peso == "*":
                    break

                # bucle para revisar la estatura
                while True:
                    estatura = input(
                        "\nIngresa la estatura del paciente en centimetros: "
                    )
                    if estatura == "*":
                        break
                    if estatura == "":
                        print("No puedes OMITIR este valor. Intenta de nuevo.")
                        continue
                    if estatura.isalpha():
                        print(
                            "Estas intentando introducir valores de TEXTO. Intenta de nuevo."
                        )
                        continue
                    try:
                        estatura = float(estatura)
                    except Exception:
                        print("Porfavor introduce valores NUMERICOS. Intenta de nuevo.")
                        continue
                    if estatura > 0:
                        # rompe el bucle para revisar la estatura
                        break
                    else:
                        print(
                            "No puedes ingresar una estatura NEGATIVA. Intenta de nuevo."
                        )

                if estatura == "*":
                    break

                # bucle para la revision de la presion en general sistolica y asistolica
                # bucle para revisar la presion sistolca 120
                while True:
                    presion_sistolica_120 = input(
                        "\nIngrese el valor de la presion Sistolica: "
                    )
                    if presion_sistolica_120 == "*":
                        break
                    if presion_sistolica_120 == "":
                        print("No puedes omitir este campo. Intenta de nuevo.")
                        continue
                    try:
                        presion_sistolica_120 = int(presion_sistolica_120)
                    except ValueError:
                        print(
                            "Debes proporcionar la presion 'sistolica' con caracteres numericos. Intenta de nuevo."
                        )
                        continue
                    if presion_sistolica_120 > 0:
                        # rompe el bucle para revisar la presion sistolca 120
                        break
                    else:
                        print(
                            "Debes proporcionar la presion 'sistolica' con numeros enternos positivos."
                        )
                        continue
                if presion_sistolica_120 == "*":
                    break

                # bucle para revsar la presion asistolica 80
                while True:
                    presion_asistolica_80 = input(
                        "\nIngrese el valor de la presion Asistolica : "
                    )
                    if presion_asistolica_80 == "*":
                        break
                    if presion_asistolica_80 == "":
                        print("No puedes omitir este campo. Intenta de nuevo.")
                        continue
                    try:
                        presion_asistolica_80 = int(presion_asistolica_80)
                    except ValueError:
                        print(
                            "Debes proporcionar la presion 'asistolica' con caracteres numericos. Intenta de nuevo."
                        )
                        continue

                    if presion_asistolica_80 > 0:
                        # rompe el bucle para revsar la presion asistolica 80
                        break
                    else:
                        print(
                            "Debes proporcionar la presion 'asistolica' con numeros enternos positivos."
                        )
                        continue
                if presion_asistolica_80 == "*":
                    break

                if (
                    0 <= presion_sistolica_120 <= 999
                    and 0 <= presion_asistolica_80 <= 999
                ):
                    # convirtiendo la persion sistolica y asistolica en un folo formato
                    presion_arterial = (
                        f"{presion_sistolica_120:03d}/{presion_asistolica_80:03d}"
                    )
                    # rompe el bucle para la revision de la presion en general sistolica y asistolica
                else:
                    print("La presion debe estar en un rango de entre 0-999")
                    continue

                # bucle para la revision del diagnostico
                while True:
                    diagnostico = input("\nDiagnostico: \n\t==>")
                    if diagnostico == "*":
                        break
                    if diagnostico == "":
                        print("No se puede omitir este campo. Intente de nuevo.")
                        continue

                    medicion_longitud_sin_espacios_blanco = len(
                        diagnostico.replace(" ", "")
                    )

                    if medicion_longitud_sin_espacios_blanco > 200:
                        print(
                            "\nEl dianostico supera la longitud maxima de 200 caracteres."
                        )
                        print(
                            f"Despues de superar la longitud maxima de 200 caracteres el diagnostico se muestra de esta forma: \n\t{diagnostico[:200]}"
                        )
                        decision_guardar = input(
                            "Deseas guardar de esta forma el diagnostico (SI/NO): "
                        ).upper()
                        if decision_guardar.upper() == "SI":
                            diagnostico = diagnostico[:200]
                            break
                        else:
                            continue
                    else:
                        break

                if diagnostico == "*":
                    break

                # creeando la conexion para extraerle la fecha_nacimiento de un determinado paciente, dandole el atributo de fechas
                try:
                    with sqlite3.connect("base_datos_ev3.db") as conn:
                        cursor = conn.cursor()
                        cursor.execute(
                            "SELECT r.fecha_nacimiento FROM registro_pacientes r\
                            INNER JOIN citas c ON c.id_paciente = r.id_paciente WHERE c.id_folio = ?",
                            (id_folio_realizar_cita,),
                        )

                        resultado = cursor.fetchall()

                        # accediendo a el primer valore de la lista, a el primer elemento de la tupla
                        fecha_nacimiento = resultado[0][0]

                        fecha_nacimiento = datetime.strptime(
                            fecha_nacimiento, "%Y-%m-%d"
                        ).date()

                except sqlite3.Error as e:
                    print(e)
                finally:
                    if conn:
                        conn.close()

                # sacando la hora de llegada del paciente
                hora_llegada = datetime.now().strftime("%H:%M")

                # sacando la fecha actual para calcular su edad del paciente
                fecha_actual = datetime.today().date()

                fecha_actual = datetime.today().date()

                # se restan los años y despues se compara si ya se han pasado el dia de su cumpleaños
                edad = (
                    fecha_actual.year
                    - fecha_nacimiento.year
                    - (
                        (fecha_actual.month, fecha_actual.day)
                        < (fecha_nacimiento.month, fecha_nacimiento.day)
                    )
                )

                # datos que se actualizaran
                datos = (
                    hora_llegada,
                    peso,
                    estatura,
                    presion_arterial,
                    diagnostico,
                    edad,
                    id_folio_realizar_cita,
                )

                try:
                    with sqlite3.connect("base_datos_ev3.db") as conn:
                        cursor = conn.cursor()
                        # seleccionando la hora_llegada para saber si tiene la cita realizada o se le puede realizar
                        cursor.execute(
                            "SELECT hora_llegada FROM citas WHERE id_folio = ?;",
                            (id_folio_realizar_cita,),
                        )

                        # contendra "None" si la cita ya se ha realizado
                        validando_cita = cursor.fetchall()

                        # validando que esta cita no se haya realizado
                        if validando_cita[0][0] == None:
                            cursor.execute(
                                "UPDATE citas \
                            SET hora_llegada = ?, peso = ?, estatura = ?, presion_arterial = ?, diagnostico = ?, edad = ? \
                            WHERE id_folio = ?",
                                datos,
                            )
                        # la cita ya contiene datos
                        else:
                            print(
                                "Este paciente ya ha realizado su cita. Intente con otro."
                            )

                except sqlite3.Error as e:
                    print(e)
                finally:
                    if conn:
                        conn.close()

                # rompe el bucle
                break
            else:
                print(f"Folio {id_folio_realizar_cita} no encontrado")

        else:
            print("No existen citas programadas a las cuales realizarle la cita")


# def listo
def cancelar_cita_busqueda_por_fecha():
    # bucle que recibe la fecha y la compara
    while True:
        print("\nAVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")

        while True:
            fecha_ingresada = input(
                "\nIngrese una fecha que desea buscar en el formato (MM/DD/YYYY): "
            )
            if fecha_ingresada == "*":
                break
            if fecha_ingresada == "":
                print("No puedes omitir este campo. Intenta de nuevo.")
                continue
            try:
                fecha_ingresada = datetime.strptime(fecha_ingresada, "%m/%d/%Y").date()
                break
            except ValueError:
                print("Ingresa la fecha en el formato (MM/DD/YYYY)")
                continue

        if fecha_ingresada == "*":
            break

        # sacando la fecha actual para  mostrar citas futuras
        fecha_actual = datetime.today().date()

        # recibir la fecha, despues en esas fechas ver los folios
        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT c.id_folio, r.primer_apellido, r.segundo_apellido, r.nombre, c.turno\
                    FROM citas c \
                    INNER JOIN registro_pacientes r ON r.id_paciente = c.id_paciente \
                    WHERE c.fecha_cita = ? AND c.fecha_cita > ? AND c.hora_llegada IS NULL",
                    (fecha_ingresada, fecha_actual),
                )

                resultado = cursor.fetchall()

                if resultado:
                    print(
                        "============================================CITA ENCONTRADA(S)============================================"
                    )
                    print(
                        f"{'FOLO':^5}\t{'PRIMER APELLIDO':<10}\t\t{'SEGUNDO APELLIDO':<10}\t{'NOMBRE':<10}\t{'TURNO':<10}"
                    )
                    for fila in resultado:
                        (
                            folio,
                            primer_apellido,
                            segundo_apellido,
                            nombre,
                            turno,
                        ) = fila
                        print(
                            f"{folio:^5}\t{primer_apellido:<10}\t\t{segundo_apellido:<10}\t\t{nombre:<10}\t{turno:<10}"
                        )

                    # bucle validacion folio
                    while True:
                        folio_cita_cancelar = input(
                            "Ingrese el folio de la cita que desea cancelar: "
                        )
                        if folio_cita_cancelar == "*":
                            break
                        try:
                            folio_cita_cancelar = int(folio_cita_cancelar)
                            # rompe el bucle validacion folio
                            break
                        except ValueError:
                            print("Ingresa valores numericos")
                            continue
                    if folio_cita_cancelar == "*":
                        break

                    # va i guarda los folios indice 0 de todas las tuplas
                    lis_comp_folio = [folio[0] for folio in resultado]

                    # comprueba si el folio que solicite existe en la lista de folios
                    if folio_cita_cancelar in lis_comp_folio:
                        confirmacion_eliminacion_cita = input(
                            "Confirmas que deseas eliminar esta cita (SI/NO): "
                        ).upper()
                        if confirmacion_eliminacion_cita == "SI":
                            cursor.execute(
                                "DELETE FROM citas WHERE id_folio = ?",
                                (folio_cita_cancelar,),
                            )
                            print(f"Cita {folio_cita_cancelar} eliminada con exito")
                            break
                    else:
                        print(
                            f"El folio proporcionadao {folio_cita_cancelar} no existe"
                        )

                else:
                    print(
                        "La fecha seleccionada no comple con las condiciones. No se tiene que haber llevado a cabo y debe ser una cita futura"
                    )

        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()


# def listo
def cancelar_cita_busqueda_por_paciente():
    # bucle cancelado la cita por folio
    while True:
        print("AVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")
        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()

                # select que muestra los datos del paciente con su folio donde su cita no se ha realizado
                cursor.execute(
                    "SELECT r.id_paciente, r.primer_apellido, r.segundo_apellido, r.nombre\
                    FROM citas c \
                    INNER JOIN registro_pacientes r ON r.id_paciente = c.id_paciente \
                    WHERE c.hora_llegada IS NULL\
                    GROUP BY r.id_paciente"
                )

                resultado = cursor.fetchall()

                if resultado:
                    print(
                        "====================================PACIENTES CON CITAS PENDIENTES===================================="
                    )
                    print(
                        f"{'CLAVE':^5}\t{'PRIMER APELLIDO':<10}\t\t{'SEGUNDO APELLIDO':<10}\t{'NOMBRE':<10}"
                    )
                    for fila in resultado:
                        (
                            clave,
                            primer_apellido,
                            segundo_apellido,
                            nombre,
                        ) = fila
                        print(
                            f"{clave:^5}\t{primer_apellido:<10}\t\t{segundo_apellido:<10}\t\t{nombre:<10}"
                        )

                    # bucle comprobando la cleve del paciente
                    while True:
                        clave_paciente_seleciconado = input(
                            "Ingresa la clave del paciente deseado: "
                        )
                        if clave_paciente_seleciconado == "*":
                            break
                        try:
                            clave_paciente_seleciconado = int(
                                clave_paciente_seleciconado
                            )
                            # rompe el bucle comprobando la cleve del paciente
                            break
                        except ValueError:
                            print("Ingresa valores numericos")
                    if clave_paciente_seleciconado == "*":
                        break

                    # leyendo la clave del paciente

                    list_comp_claves = [clave[0] for clave in resultado]

                    if clave_paciente_seleciconado in list_comp_claves:

                        # sacando la fecha actual para  mostrar citas futuras
                        fecha_actual = datetime.today().date()

                        cursor.execute(
                            "SELECT c.id_folio, c.fecha_cita, c.turno \
                            FROM citas c \
                            INNER JOIN registro_pacientes r ON r.id_paciente = c.id_paciente \
                            WHERE r.id_paciente = ? AND c.fecha_cita > ? AND c.hora_llegada IS NULL",
                            (
                                clave_paciente_seleciconado,
                                fecha_actual,
                            ),
                        )
                        resultado = cursor.fetchall()

                        if resultado:
                            print(
                                f"\n\n========CITAS DEL PACIENTE SELECCIONADO {clave_paciente_seleciconado}========"
                            )
                            print(f"{'FOLIO':^5}\t{'FECHA CITA':<10}\t{'TURNO':<10}")
                            for (
                                folio,
                                fecha_cita,
                                turno,
                            ) in resultado:
                                print(f"{folio:^5}\t{fecha_cita:<10}\t{turno:<10}")

                            # bucle para solicitar folio a cancelar
                            while True:
                                print(
                                    "\nAVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'"
                                )
                                id_folio_cancelar = input(
                                    "\nIngresa el folio de la cita que desea cancelar: "
                                )
                                if id_folio_cancelar == "*":
                                    break
                                if id_folio_cancelar == "":
                                    continue
                                try:
                                    id_folio_cancelar = int(id_folio_cancelar)
                                    # rompe el bucle para solicitar folio a cancelar
                                    break
                                except ValueError:
                                    print("Ingresa numeros enteros")
                            if id_folio_cancelar == "*":
                                break

                            # list comprehension de los folos para validar que el folio que me mande exista
                            lis_comp_folio = [folio[0] for folio in resultado]

                            if id_folio_cancelar in lis_comp_folio:
                                while True:
                                    confirmacion_folio_cancelar = input(
                                        f"Confirmas que deseas cancelar la cita con el folio {id_folio_cancelar} (SI/NO): "
                                    ).upper()
                                    if confirmacion_folio_cancelar == "SI":
                                        cursor.execute(
                                            "DELETE FROM citas WHERE id_folio = ?",
                                            (id_folio_cancelar,),
                                        )
                                    # rompe el bucle de la confirmacion para la cancelacion de la cita
                                    break
                            else:
                                print("El folio seleccionado no existe")

                        else:
                            print(
                                "El paciente seleccionado no seleccionado no tiene citas programadas"
                            )
                    else:
                        print("El paciente con la clave seleccionada no existe")
                else:
                    print("No existen citas pendientes por realizar")

        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()
        # romiendo el bucle cancelado la cita por folio
        break


# def listo
def sub_menu_cancelacion_citas():
    try:
        with sqlite3.connect("base_datos_ev3.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id_folio FROM citas")
            resultado = cursor.fetchall()
    except sqlite3.Error as e:
        print(e)

    # bucle del menu principal
    while True:
        if resultado:
            print("\n\n================SUB MENU CANCELACION DE CITAS================")
            print(
                "1. Busqueda por fecha \n2. Busqueda por paciente. \n3. Volver al menu anterior. "
            )
            print("===============================================================")

            opcion_sub_menu_cancelacion_de_citas = input("Ingrese la opcion deseada: ")
            try:
                opcion_sub_menu_cancelacion_de_citas = int(
                    opcion_sub_menu_cancelacion_de_citas
                )
            except ValueError:
                print("Ingresa un valor numerico. Intenta de nuevo.")
                continue

            # cancelacion de busqueda por fecha
            if opcion_sub_menu_cancelacion_de_citas == 1:
                cancelar_cita_busqueda_por_fecha()
            elif opcion_sub_menu_cancelacion_de_citas == 2:
                cancelar_cita_busqueda_por_paciente()
            elif opcion_sub_menu_cancelacion_de_citas == 3:
                break
        else:
            print("No existen citas programadas a las cuales cancelar.")
            # rompe el bucle del menu principal
            break


# Bloque principal del programa
# ya tiene el registro de los pacientes
def bloque_principal():
    creacion_tablas()
    while True:
        limpiar_consola()
        print("\n\n==============MENU PRINCIPAL==============")
        print(
            "1. Registro de pacientes. \n2. Citas. \n3. Consultas y reportes. \n4. Salir del sistema."
        )
        print("==========================================")
        while True:
            opcion_menu_principal = input("Ingresa la opcion deseada: ")
            if opcion_menu_principal == "":
                print("No puedes OMITIR este campo. Intenta de nuevo.")
                continue
            try:
                opcion_menu_principal = int(opcion_menu_principal)
                break
            except Exception:
                print("Porfavor introduce valores NUMERICOS. Intenta de nuevo.")
                continue

        # registrar pacientes
        if opcion_menu_principal == 1:
            registrar_pacientes()

        # citas
        elif opcion_menu_principal == 2:
            while True:
                print("\n================SUBMENU CITAS================")
                print(
                    "1. Programacion de citas. \n2. Realizacion de citas programadas. \n3. Cancelacion de citas. \n4. Volver al menu anterior."
                )
                print("===============================================")
                opcion_sub_menu_citas = input("Ingresa la opcion deseada: ")

                try:
                    opcion_sub_menu_citas = int(opcion_sub_menu_citas)
                except Exception:
                    print("Ingresa un valor NUMERICO. Intenta de nuevo")

                if opcion_sub_menu_citas in [1, 2, 3, 4]:
                    pass
                else:
                    print("Debes ingresar una opcion entre 1-4")
                    continue

                if opcion_sub_menu_citas == 1:
                    programar_citas()
                elif opcion_sub_menu_citas == 2:
                    realizar_cita()
                elif opcion_sub_menu_citas == 3:
                    sub_menu_cancelacion_citas()
                elif opcion_sub_menu_citas == 4:
                    break

        # consultas y resportes
        elif opcion_menu_principal == 3:
            sub_menu_consultas_y_reportes()

        # salir del sistema
        elif opcion_menu_principal == 4:
            salir = input("Confirmas que deseas salir del sistema (SI/NO): ").upper()
            if salir == "SI":
                print("Saliendo del sistema...")
                break


def sub_menu_consultas_y_reportes():
    while True:
        print("\n\n============SUB MENU CONSULTAS Y REPORTES============")
        print(
            "1. Reportes de citas. \n2. Reportes de pacientes. \n3. Estadisticos demograficos. \n4. Volver al menu anterior."
        )
        print("=====================================================")
        opcion_consultas_y_reportes = input("Seleccione la opcion deseada: ")
        if opcion_consultas_y_reportes == "":
            print("No puedes OMITIR este campo. Intenta de nuevo.")
            continue
        try:
            opcion_consultas_y_reportes = int(opcion_consultas_y_reportes)

        except Exception:
            print("Porfavor introduce valores enteros. Intenta de nuevo.")
            continue

        # reportes de citas
        if opcion_consultas_y_reportes == 1:
            # def lista
            reportes_de_citas()

        # reportes de pacientes
        elif opcion_consultas_y_reportes == 2:
            reportes_de_pacientes()

        # estadisticos demograficos
        elif opcion_consultas_y_reportes == 3:
            estadisticos_demograficos()

        # volver al menu anterior
        elif opcion_consultas_y_reportes == 4:
            break

        else:
            print("Selecciona entre las opciones 1-4")
            continue


# def lista
def reportes_de_citas():
    try:
        with sqlite3.connect("base_datos_ev3.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id_folio FROM citas")

            resultado = cursor.fetchall()

            if resultado:
                while True:
                    print("\n\n===========REPORTES DE CITAS==========")
                    print(
                        "1. Por periodo. \n2. Por paciente. \n3. Regresar al menu de consultas y reportes."
                    )
                    print("======================================")
                    opcion_reportes_de_citas = input("Seleccione la opcion deseada: ")
                    if opcion_reportes_de_citas == "":
                        print("No puedes OMITIR este campo. Intenta de nuevo.")
                        continue
                    try:
                        opcion_reportes_de_citas = int(opcion_reportes_de_citas)
                    except Exception:
                        print("Porfavor introduce valores enteros. Intenta de nuevo.")
                        continue

                    # 1. Por periodo
                    if opcion_reportes_de_citas == 1:
                        reporte_de_citas_por_periodo()
                    # 2. Por paciente.
                    elif opcion_reportes_de_citas == 2:
                        # reporte_de_citas_por_paciente()
                        reporte_de_citas_por_paciente_new_version()

                    # \n3. Regresar al menu de consultas y reportes.
                    elif opcion_reportes_de_citas == 3:
                        break
                    else:
                        print(
                            "Porfavor selecciona un numero entre el 1-3. Intenta de nuevo."
                        )
                        continue
            else:
                print("No existen citas")
    except sqlite3.Error as e:
        print(e)
    finally:
        if conn:
            conn.close()


# def lista
def reporte_de_citas_por_periodo():
    while True:
        print("\nAVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")

        while True:
            fecha_inicio = input("\nIngrese la fecha inicial del periodo (MM/DD/YYYY):")
            if fecha_inicio == "*":
                break
            if fecha_inicio == "":
                print("No puedes OMITIR este campo. Intenta de nuevo.")
                continue
            if fecha_inicio.isalpha():
                print("Estas intentando introducir valores de TEXTO. Intenta de nuevo.")
                continue
            if fecha_inicio.isdigit():
                print("Estas intentando introducir valores ENTEROS. Intenta de nuevo.")
                continue
            try:
                fecha_inicio = datetime.strptime(fecha_inicio, "%m/%d/%Y").date()
                break
            except Exception:
                print(
                    "Ingresa la fecha en el formato valido (MM/DD/YYYY). Intenta de nuevo."
                )
                continue

        if fecha_inicio == "*":
            break

        while True:
            fecha_fin = input("\nIngrese la fecha inicial del periodo (MM/DD/YYYY):")
            if fecha_fin == "*":
                break
            if fecha_fin == "":
                print("No puedes OMITIR este campo. Intenta de nuevo.")
                continue
            if fecha_fin.isalpha():
                print("Estas intentando introducir valores de TEXTO. Intenta de nuevo.")
                continue
            if fecha_fin.isdigit():
                print("Estas intentando introducir valores ENTEROS. Intenta de nuevo.")
                continue
            try:
                fecha_fin = datetime.strptime(fecha_fin, "%m/%d/%Y").date()
                break
            except Exception:
                print(
                    "Ingresa la fecha en el formato valido (MM/DD/YYYY). Intenta de nuevo."
                )
                continue

        if fecha_fin == "*":
            break

        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT c.id_folio, r.primer_apellido, r.segundo_apellido, r.nombre, \
                    c.fecha_cita, c.hora_llegada, c.peso, c.estatura, c.presion_arterial, c.edad \
                    FROM citas c \
                    INNER JOIN registro_pacientes r ON r.id_paciente = c.id_paciente \
                    WHERE c.fecha_cita BETWEEN ? AND  ?;",
                    (fecha_inicio, fecha_fin),
                )

                resultado = cursor.fetchall()
                if resultado:
                    print(
                        f"=================================================CITAS ENCONTRADAS ENTRE {fecha_inicio} AL {fecha_fin}================================================="
                    )
                    print(
                        f"{'FOLIO':^5}\t{'PRIMER APELLIDO':<10}\t\t{'SEGUNDO APELLIDO':<10}\t{'NOMBRE':<10}\t{'FECHA CITA':<10}\t{'HORA LLEGADA':<10}\t{'PESO':<10}\t{'ESTATURA':<10}\t{'PRESION ARTERIAL':<10}\t{'EDAD':<10}"
                    )
                    for tupla in resultado:
                        (
                            folio,
                            primer_apellido,
                            segundo_apellido,
                            nombre,
                            fecha_cita,
                            hora_llegada,
                            peso,
                            estatura,
                            presion_arterial,
                            edad,
                        ) = tupla
                        print(
                            f"{folio:^5}\t{primer_apellido:<10}\t\t{segundo_apellido:<10}\t\t{nombre:<10}\t{fecha_cita:<10}\t{hora_llegada}\t\t{peso}\t\t{estatura}\t\t{presion_arterial}\t\t\t{edad}"
                        )
                    # bucle exportando a formatos csv y exel
                    while True:
                        confirmacion_exportacion = input(
                            "\nSi deseas exportar esta informacion ingresa una de las opciones (1. CSV 2. EXEL 3. No por el momento): "
                        )
                        if confirmacion_exportacion == "*":
                            break
                        if confirmacion_exportacion == "":
                            print("No puedes omitir este valor")
                            continue
                        try:
                            confirmacion_exportacion = int(confirmacion_exportacion)
                            break
                        except ValueError:
                            print("Ingresa numeros enteros")

                    if confirmacion_exportacion == "*":
                        break

                    if confirmacion_exportacion == 1:
                        with open(
                            f"reporte_por_periodo_citas_{fecha_inicio}_al_{fecha_fin}.csv",
                            "w",
                            newline="",
                        ) as file:
                            writer = csv.writer(file)
                            encabezado = [
                                "FOLIO",
                                "PRIMER APELLIDO",
                                "SEGUNDO APELLIDO",
                                "NOMBRE",
                                "FECHA CITA",
                                "HORA LLEGADA",
                                "PESO",
                                "ESTATURA",
                                "PRESION ARTERIAL",
                                "EDAD",
                            ]

                            writer.writerow(encabezado)
                            writer.writerows(resultado)
                            print("Exportado correctamente en formato 'csv'")
                            break

                    if confirmacion_exportacion == 2:
                        libro = openpyxl.Workbook()
                        hoja = libro.active

                        encabezado = [
                            "FOLIO",
                            "PRIMER APELLIDO",
                            "SEGUNDO APELLIDO",
                            "NOMBRE",
                            "FECHA CITA",
                            "HORA LLEGADA",
                            "PESO",
                            "ESTATURA",
                            "PRESION ARTERIAL",
                            "EDAD",
                        ]

                        hoja.append(encabezado)

                        for fila in resultado:
                            hoja.append(fila)

                        libro.save(
                            f"reporte_por_periodo_citas_{fecha_inicio}_al_{fecha_fin}.xlsx"
                        )
                        print("Exportado correctamente en formato 'xlsx'")
                        break

                    if confirmacion_exportacion == 3:
                        break
                    else:
                        print("Selecciona una opcion entre 1-3")
                        continue

                else:
                    print("No se encontraron citas en este periodo")

        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()

        # rompe el bucle para el aviso
        break


# def controverisa
def reporte_de_citas_por_paciente():
    # bucle de aviso
    while True:
        print("AVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")

        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT c.id_folio FROM citas c INNER JOIN registro_pacientes r ON r.id_paciente = c.id_paciente"
                )

                resultado_folios = cursor.fetchall()

                if resultado_folios:
                    # creeando una list comprehension que guarda los folios
                    list_comp_folios = [folio[0] for folio in resultado_folios]

                    print("====FOLIOS DISPONIBLES====")
                    for folio in list_comp_folios:
                        print(f"\t{folio:^5}")

                    while True:
                        id_folio_buscar = input("\nIngresa el folio de la cita: ")
                        if id_folio_buscar == "*":
                            break
                        if id_folio_buscar == "":
                            print("No puedes omitir este campo")
                            continue
                        try:
                            id_folio_buscar = int(id_folio_buscar)
                            # rompe el bucle validacoin del id del paciente
                            break
                        except ValueError:
                            print("Ingresa valores numericos")

                    if id_folio_buscar == "*":
                        break

                    if id_folio_buscar in list_comp_folios:
                        cursor.execute(
                            "SELECT c.id_folio, r.primer_apellido, r.segundo_apellido, r.nombre, c.fecha_cita, c.hora_llegada, \
                            c.peso, c.estatura, c.presion_arterial, c.edad \
                            FROM citas c \
                            INNER JOIN registro_pacientes r ON r.id_paciente = c.id_paciente \
                            WHERE c.id_folio = ?",
                            (id_folio_buscar,),
                        )

                        resultado = cursor.fetchall()

                        if resultado:
                            print(
                                f"==========================================================================CITA DEL PACIENTE=========================================================================="
                            )
                            print(
                                f"{'FOLIO':^5}\t{'PRIMER APELLIDO':<10}\t\t{'SEGUNDO APELLIDO':<10}\t{'NOMBRE':<10}\t{'FECHA CITA':<10}\t{'HORA LLEGADA':<10}\t{'PESO':<10}\t{'ESTATURA':<10}\t{'PRESION ARTERIAL':<10}\t{'EDAD':<10}"
                            )
                            for tupla in resultado:
                                (
                                    folio,
                                    primer_apellido,
                                    segundo_apellido,
                                    nombre,
                                    fecha_cita,
                                    hora_llegada,
                                    peso,
                                    estatura,
                                    presion_arterial,
                                    edad,
                                ) = tupla
                                print(
                                    f"{folio:^5}\t{primer_apellido:<10}\t\t{segundo_apellido:<10}\t\t{nombre:<10}\t{fecha_cita:<10}\t{hora_llegada}\t\t{peso}\t\t{estatura}\t\t{presion_arterial}\t\t\t{edad}"
                                )
                            #
                            while True:
                                confirmacion_exportacion = input(
                                    "\nSi deseas exportar esta informacion ingresa una de las opciones (1. CSV 2. EXEL 3. No por el momento): "
                                )
                                if confirmacion_exportacion == "*":
                                    break
                                if confirmacion_exportacion == "":
                                    print("No puedes omitir este valor")
                                    continue
                                try:
                                    confirmacion_exportacion = int(
                                        confirmacion_exportacion
                                    )
                                    break
                                except ValueError:
                                    print("Ingresa numeros enteros")

                            if confirmacion_exportacion == "*":
                                break

                            if confirmacion_exportacion == 1:
                                with open(
                                    f"reporte_por_paciente_{id_folio_buscar}_citas.csv",
                                    "w",
                                    newline="",
                                ) as file:
                                    writer = csv.writer(file)
                                    encabezado = [
                                        "FOLIO",
                                        "PRIMER APELLIDO",
                                        "SEGUNDO APELLIDO",
                                        "NOMBRE",
                                        "FECHA CITA",
                                        "HORA LLEGADA",
                                        "PESO",
                                        "ESTATURA",
                                        "PRESION ARTERIAL",
                                        "EDAD",
                                    ]
                                    writer.writerow(encabezado)
                                    writer.writerows(resultado)
                                    print("Exportado correctamente en formato 'csv'")
                                    break

                            if confirmacion_exportacion == 2:
                                libro = openpyxl.Workbook()
                                hoja = libro.active

                                encabezado = [
                                    "FOLIO",
                                    "PRIMER APELLIDO",
                                    "SEGUNDO APELLIDO",
                                    "NOMBRE",
                                    "FECHA CITA",
                                    "HORA LLEGADA",
                                    "PESO",
                                    "ESTATURA",
                                    "PRESION ARTERIAL",
                                    "EDAD",
                                ]

                                hoja.append(encabezado)

                                for fila in resultado:
                                    hoja.append(fila)

                                libro.save(
                                    f"reporte_por_paciente_{id_folio_buscar}_citas.xlsx"
                                )
                                print("Exportado correctamente en formato 'xlsx'")
                                break

                            if confirmacion_exportacion == 3:
                                break
                            else:
                                print("Selecciona una opcion entre 1-3")
                                continue

                    else:
                        print("El folio seleccionado no existe")

                else:
                    print("No existen citas programadas")

        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()
        # rompe el bucle del aviso
        break


def reporte_de_citas_por_paciente_new_version():
    while True:
        print("\nAVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'\n")
        # mostrar los pacientes
        # seleccionar un paciente
        # si dicho paciente tiene citas mostrar la cita si la cita ya se realizo mostrar todos los datos de la cita y si no mostrar como none
        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT id_paciente, primer_apellido, segundo_apellido, nombre \
                FROM registro_pacientes"
                )

                resultado = cursor.fetchall()

                if resultado:
                    print(
                        f"{'CLAVE':^5}\t{'PRIMER APELLIDO':<10}\t\t{'SEGUNDO APELLIDO':<10}\t{'NOMBRE':<10}"
                    )
                    for fila in resultado:
                        id_paciente, primer_apellido, segundo_apellido, nombre = fila
                        print(
                            f"{id_paciente:^5}\t{primer_apellido:<10}\t\t{segundo_apellido:<10}\t\t{nombre:<10}"
                        )
                    while True:
                        clave = input("\nSelecciona un paciente: ")
                        if clave == "*":
                            break
                        if clave == "":
                            print("No puedes omitir este campo")
                            continue
                        try:
                            clave = int(clave)
                            break
                        except ValueError:
                            print("Ingresa valores enteros")

                    if clave == "*":
                        break

                    list_comp_claves = [clave[0] for clave in resultado]

                    if clave in list_comp_claves:
                        cursor.execute(
                            "SELECT c.id_folio, c.fecha_cita, c.turno, c.hora_llegada, c.peso, c.estatura, c.presion_arterial, c.diagnostico, c.edad \
                            FROM citas c \
                            INNER JOIN registro_pacientes r ON r.id_paciente = c.id_paciente \
                            WHERE c.id_paciente = ?",
                            (clave,),
                        )

                        resultado_cita = cursor.fetchall()

                        if resultado_cita:
                            print(
                                f"=========================================CITA(S) DEL PACIENTE {clave}==========================================="
                            )
                            print(
                                f"{'FOLIO':^5}\t{'FECHA CITA':<10}\t{'TURNO'}\t{'HORA LLEGADA':<10}\t{'PESO':<10}\t{'ESTATURA':<10}\t{'PRESION ARTERIAL':<10}\t{'EDAD':<10}"
                            )
                            for fila in resultado_cita:
                                (
                                    folio,
                                    fecha_cita,
                                    turno,
                                    hora_llegada,
                                    peso,
                                    estatura,
                                    presion_arterial,
                                    diagnostico,
                                    edad,
                                ) = fila
                                print(
                                    f"{folio:^5}\t{fecha_cita:<10}\t{turno}\t{hora_llegada}\t\t{peso}\t\t{estatura}\t\t{presion_arterial}\t\t\t{edad}"
                                )
                                print(f"Diagnostico: \n\t{diagnostico}\n")
                        else:
                            print(f"El paciente seleccionado {clave} no tiene citas")

                    else:
                        print(f"El paciente con la clave {clave} no existe")

                else:
                    print("No existen pacientes registrados")
        except sqlite3.Error as e:
            print("e")
        finally:
            if conn:
                conn.close()

        break


# funcion de reportes de citas por pacientex


# ===========


# bloque del submenu de reportes por pacientes
def reportes_de_pacientes():
    while True:
        print("\n\n================REPORTES DE PACIENTES================")
        print(
            "1. Listado completo de pacientes. \n2. Busqueda por clave de paciente. \n3. Busqueda por apellidos y nombres. \n4. Regresar al menu de consultas y reportes. "
        )
        print("=====================================================")
        opcion_reportes_de_pacientes = input("Seleccione la opcion deseada: ")
        if opcion_reportes_de_pacientes == "":
            print("No puedes OMITIR este campo. Intenta de nuevo.")
            continue
        try:
            opcion_reportes_de_pacientes = int(opcion_reportes_de_pacientes)
        except Exception:
            print("Porfavor introduce valores enteros. Intenta de nuevo.")
            continue

        # 1. Listado completo de pacientes.
        if opcion_reportes_de_pacientes == 1:
            listado_completo_pacientes()

        # 2. Busqueda por clave de paciente
        elif opcion_reportes_de_pacientes == 2:
            busqueda_por_clave_paciente_new_version()

        # 3. Busqueda por apellidos y nombres.
        elif opcion_reportes_de_pacientes == 3:
            busqueda_por_apellidos_y_nombres()

        # 4. Regresar al menu de consultas y reportes.
        elif opcion_reportes_de_pacientes == 4:
            break

        # opcion no valida
        else:
            print("Selecciona una opcion entre el 1-4")
            continue


# def lista
def listado_completo_pacientes():
    while True:
        print("AVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")

        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT id_paciente, primer_apellido, segundo_apellido, nombre, fecha_nacimiento, sexo FROM registro_pacientes"
                )

                resultado = cursor.fetchall()

                if resultado:
                    print(
                        "\n\n====================================LISTADO COMPLETO DE PACIENTES======================================="
                    )
                    print(
                        f"{'CLAVE':^5}\t{'PRIMER APELLIDO':<10}\t\t{'SEGUNDO APELLIDO':<10}\t{'NOMBRE':<10}\t{'FECHA NACIMIENTO':<10}\t{'SEXO':<10}"
                    )
                    for tupla in resultado:
                        (
                            clave,
                            primer_apellido,
                            segundo_apellido,
                            nombre,
                            fecha_nacimiento,
                            sexo,
                        ) = tupla
                        print(
                            f"{clave:^5}\t{primer_apellido:<10}\t\t{segundo_apellido:<10}\t\t{nombre:<10}\t{fecha_nacimiento:<10}\t\t{sexo:<10}"
                        )

                    while True:
                        confirmacion_exportacion = input(
                            "\nSi deseas exportar esta informacion ingresa una de las opciones (1. CSV 2. EXEL 3. No por el momento): "
                        )
                        if confirmacion_exportacion == "*":
                            break
                        if confirmacion_exportacion == "":
                            print("No puedes omitir este valor")
                            continue
                        try:
                            confirmacion_exportacion = int(confirmacion_exportacion)
                            break
                        except ValueError:
                            print("Ingresa numeros enteros")

                    if confirmacion_exportacion == "*":
                        break

                    if confirmacion_exportacion == 1:
                        with open(
                            "reporte_de_pacientes_listado_completo.csv",
                            "w",
                            encoding="UTF-8",
                            newline="",
                        ) as file:
                            writer = csv.writer(file)
                            encabezado = [
                                "CLAVE",
                                "PRIMER APELLIDO",
                                "SEGUNDO APELLIDO",
                                "NOMBRE",
                                "FECHA DE NACIMIENTO",
                                "SEXO",
                            ]
                            writer.writerow(encabezado)
                            writer.writerows(resultado)
                            print("Exportado correctamente en formato 'csv'")
                            break

                    if confirmacion_exportacion == 2:
                        libro = openpyxl.Workbook()
                        hoja = libro.active

                        encabezado = [
                            "CLAVE",
                            "PRIMER APELLIDO",
                            "SEGUNDO APELLIDO",
                            "NOMBRE",
                            "FECHA DE NACIMIENTO",
                            "SEXO",
                        ]

                        hoja.append(encabezado)

                        for fila in resultado:
                            hoja.append(fila)

                        libro.save("reporte_de_pacientes_listado_completo.xlsx")
                        print("Exportado correctamente en formato 'xlsx'")
                        break

                    if confirmacion_exportacion == 3:
                        break
                    else:
                        print("Selecciona una opcion entre 1-3")
                        continue

                else:
                    print("No existen pacientes registrados")

        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()

        # rompe el bucle del aviso
        break


def busqueda_por_clave_paciente_new_version():
    while True:
        print("AVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")
        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id_paciente FROM registro_pacientes")

                resultado = cursor.fetchall()

                if resultado:

                    list_comp_claves = [clave[0] for clave in resultado]
                    print("\n\n===CLAVES EXISTENTES===")
                    for clave in list_comp_claves:
                        print(f"\t{clave:^5}")

                    while True:
                        clave_buscar = input("Ingrese la clave que esta buscando: ")
                        if clave_buscar == "*":
                            break
                        if clave_buscar == "":
                            print("No puedes OMITIR este valor. Intente de nuevo.")
                            continue
                        try:
                            clave_buscar = int(clave_buscar)
                            break
                        except Exception:
                            print(
                                "Porfavor ingresa valores NUMERICOS. Intenta de nuevo."
                            )
                            continue
                    if clave_buscar == "*":
                        break

                    if clave_buscar in list_comp_claves:
                        cursor.execute(
                            "SELECT id_paciente, primer_apellido, segundo_apellido, nombre, fecha_nacimiento, sexo \
                            FROM registro_pacientes \
                            WHERE id_paciente = ?",
                            (clave_buscar,),
                        )
                        resultado_paciente_seleccionado = cursor.fetchall()
                        print(
                            f"\n\n================================INFORMACION DEL PACIENTE SELECCIONADO {clave_buscar}================================"
                        )
                        print(
                            f"{'CLAVE':^5}\t{'PRIMER APELLIDO':<10}\t\t{'SEGUNDO APELLIDO':<10}\t{'NOMBRE':<10}\t{'FECHA NACIMIENTO':<10}\t{'SEXO':<10}"
                        )
                        for (
                            id_paciente,
                            primer_apellido,
                            segundo_apellido,
                            nombre,
                            fecha_nacimiento,
                            sexo,
                        ) in resultado_paciente_seleccionado:
                            print(
                                f"{id_paciente:^5}\t{primer_apellido:<10}\t\t{segundo_apellido:<10}\t\t{nombre:<10}\t{fecha_nacimiento:<10}\t\t{sexo:<10}"
                            )
                        #
                        while True:
                            confirmacion_exportacion = input(
                                "\nSi deseas exportar esta informacion ingresa una de las opciones (1. CSV 2. EXEL 3. No por el momento): "
                            )
                            if confirmacion_exportacion == "*":
                                break
                            if confirmacion_exportacion == "":
                                print("No puedes omitir este valor")
                                continue
                            try:
                                confirmacion_exportacion = int(confirmacion_exportacion)
                                break
                            except ValueError:
                                print("Ingresa numeros enteros")
                        if confirmacion_exportacion == "*":
                            break

                        if confirmacion_exportacion == 1:
                            with open(
                                f"reporte_paciente_seleccionado_{clave_buscar}.csv",
                                "w",
                                encoding="UTF-8",
                                newline="",
                            ) as file:
                                writer = csv.writer(file)
                                encabezado = [
                                    "CLAVE",
                                    "PRIMER APELLIDO",
                                    "SEGUNDO APELLIDO",
                                    "NOMBRE",
                                    "FECHA DE NACIMIENTO",
                                    "SEXO",
                                ]
                                writer.writerow(encabezado)
                                writer.writerows(resultado_paciente_seleccionado)
                                print("Exportado correctamente en formato 'csv'")

                        elif confirmacion_exportacion == 2:
                            libro = openpyxl.Workbook()
                            hoja = libro.active

                            encabezado = [
                                "CLAVE",
                                "PRIMER APELLIDO",
                                "SEGUNDO APELLIDO",
                                "NOMBRE",
                                "FECHA DE NACIMIENTO",
                                "SEXO",
                            ]

                            hoja.append(encabezado)

                            for fila in resultado_paciente_seleccionado:
                                hoja.append(fila)

                            libro.save(
                                f"reporte_paciente_seleccionado_{clave_buscar}.xlsx"
                            )
                            print("Exportado correctamente en formato 'xlsx'")

                        elif confirmacion_exportacion == 3:
                            pass
                        else:
                            print("Selecciona una opcion entre 1-3")
                            continue

                        # buscle para saber si necesit el expediente
                        while True:
                            confirmacion_expediente = input(
                                "\nSi deseas consultar el expediente de este paciente confirma con un (SI/NO): "
                            ).upper()
                            if confirmacion_expediente == "*":
                                break
                            if confirmacion_expediente == "":
                                print("No se puede omitir este campo")
                                continue
                            if confirmacion_expediente == "SI":
                                cursor.execute(
                                    "SELECT id_folio, fecha_cita, turno, hora_llegada, peso, estatura, presion_arterial, diagnostico, edad \
                                    FROM citas \
                                    WHERE id_paciente = ? AND hora_llegada IS NOT NULL",
                                    (clave_buscar,),
                                )

                                resultado_expediente = cursor.fetchall()

                                if resultado_expediente:
                                    print(
                                        f"\n\n=======================================================EXPEDIENTE DEL PACIENTE {clave_buscar}======================================================"
                                    )
                                    print(
                                        f"\n\n{'FOLIO':^5}\t{'FECHA CITA':^5}\t{'TURNO':^5}\t{'HORA DE LLAGADA':<10}\t\t{'PESO':<10}\t{'ESTATURA':<10}\t{'PRESION ARTERIAL':^5}\t{'EDAD':<10}"
                                    )
                                    for tupla in resultado_expediente:
                                        (
                                            id_folio,
                                            fecha_cita,
                                            turno,
                                            hora_llegada,
                                            peso,
                                            estatura,
                                            presion_arterial,
                                            diagnostico,
                                            edad,
                                        ) = tupla
                                        print(
                                            f"{id_folio:^5}\t{fecha_cita:^5}\t{turno:^5}\t{hora_llegada}\t\t\t{peso}\t\t{estatura}\t\t{presion_arterial}\t\t\t{edad}"
                                        )
                                        print(f"Diagnostico: \n\t:{diagnostico}\n")
                                else:
                                    print(
                                        f"El paciente con la clave {clave_buscar} no tiene citas realizadas"
                                    )

                            else:
                                break
                            # rompe el bucle para saber si necesita el expediente
                            break

                    else:
                        print(f"La clave proporcionada {clave_buscar} no existe")

                else:
                    print("No existen pacientes registrados")
        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()

        # rompe el bucle del aiviso
        break


# def lista
def busqueda_por_apellidos_y_nombres():
    try:
        with sqlite3.connect("base_datos_ev3.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id_paciente FROM registro_pacientes")

            resultado = cursor.fetchall()

            if resultado:
                # bucle del aviso
                while True:
                    print(
                        "\n\nAVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'"
                    )

                    # bucle del apellido_buscar
                    while True:
                        apellido_buscar = input(
                            "\nIngrese el apellido del paciente que desea buscar: "
                        ).upper()
                        if apellido_buscar == "*":
                            break
                        if apellido_buscar == "":
                            print("No puedes OMITIR este valor. Intenta de nuevo.")
                            continue
                        if apellido_buscar.isdigit():
                            print(
                                "Estas intentando ingresar valores NUMERICOS. Intenta de nuevo."
                            )
                            continue
                        break

                    if apellido_buscar == "*":
                        break
                    # bucle del nombre_buscar
                    while True:
                        nombre_buscar = input(
                            "\nIngrese el nombre del paciente que desea buscar: "
                        ).upper()
                        if nombre_buscar == "*":
                            break
                        if nombre_buscar == "":
                            print("No peudes OMITIR este valor. Intenta de nuevo.")
                            continue
                        if nombre_buscar.isdigit():
                            print(
                                "Estas intentando ingresar valores NUMERICOS. Intenta de nuevo."
                            )
                            continue
                        # rompe el bucle del nombre_buscar
                        break

                    if nombre_buscar == "*":
                        break

                    cursor.execute(
                        "SELECT id_paciente, primer_apellido, segundo_apellido, nombre, fecha_nacimiento, sexo \
                        FROM registro_pacientes \
                        WHERE UPPER(primer_apellido) = ? OR UPPER(segundo_apellido) = ? OR UPPER(nombre) = ?",
                        (
                            apellido_buscar.upper(),
                            apellido_buscar.upper(),
                            nombre_buscar.upper(),
                        ),
                    )

                    resultado_pacientes_encontrados = cursor.fetchall()

                    if resultado_pacientes_encontrados:
                        print(
                            f"========================================PACIENTES ENCONTRADO========================================"
                        )
                        print(
                            f"{'CLAVE':^5}\t{'PRIMER APELLIDO':<10}\t\t{'SEGUNDO APELLIDO':<10}\t{'NOMBRE':<10}\t{'FECHA NACIMIENTO':<10}\t{'SEXO':<10}"
                        )
                        for (
                            id_paciente,
                            primer_apellido,
                            segundo_apellido,
                            nombre,
                            fecha_nacimiento,
                            sexo,
                        ) in resultado_pacientes_encontrados:
                            print(
                                f"{id_paciente:^5}\t{primer_apellido:<10}\t\t{segundo_apellido:<10}\t\t{nombre:<10}\t{fecha_nacimiento:<10}\t\t{sexo:<10}"
                            )
                        #
                        while True:
                            confirmacion_exportacion = input(
                                "\nSi deseas exportar esta informacion ingresa una de las opciones (1. CSV 2. EXEL 3. No por el momento): "
                            )
                            if confirmacion_exportacion == "*":
                                break
                            if confirmacion_exportacion == "":
                                print("No puedes omitir este valor")
                                continue
                            try:
                                confirmacion_exportacion = int(confirmacion_exportacion)
                                break
                            except ValueError:
                                print("Ingresa numeros enteros")
                        if confirmacion_exportacion == "*":
                            break

                        if confirmacion_exportacion == 1:
                            with open(
                                f"reporte_de_pacientes_encontrados_por_nombres_apellidos.csv",
                                "w",
                                encoding="UTF-8",
                                newline="",
                            ) as file:
                                writer = csv.writer(file)
                                encabezado = [
                                    "CLAVE",
                                    "PRIMER APELLIDO",
                                    "SEGUNDO APELLIDO",
                                    "NOMBRE",
                                    "FECHA DE NACIMIENTO",
                                    "SEXO",
                                ]
                                writer.writerow(encabezado)
                                writer.writerows(resultado_pacientes_encontrados)
                                print("Exportado correctamente en formato 'csv'")

                        elif confirmacion_exportacion == 2:
                            libro = openpyxl.Workbook()
                            hoja = libro.active

                            encabezado = [
                                "CLAVE",
                                "PRIMER APELLIDO",
                                "SEGUNDO APELLIDO",
                                "NOMBRE",
                                "FECHA DE NACIMIENTO",
                                "SEXO",
                            ]

                            hoja.append(encabezado)

                            for fila in resultado_pacientes_encontrados:
                                hoja.append(fila)

                            libro.save(
                                f"reporte_de_pacientes_encontrados_por_nombres_apellidos.csv.xlsx"
                            )
                            print("Exportado correctamente en formato 'xlsx'")

                        elif confirmacion_exportacion == 3:
                            pass
                        else:
                            print("Selecciona una opcion entre 1-3")
                            continue

                        # bucle de la validacion de la clave del paciente
                        while True:
                            clave_paciente_a_buscar_expediente = input(
                                "\nSi desea consultar el expediente de algun paciente ingrese la clave: "
                            )
                            if clave_paciente_a_buscar_expediente == "*":
                                break
                            if clave_paciente_a_buscar_expediente == "":
                                print("No se puede omitir este campo")
                                continue
                            try:
                                clave_paciente_a_buscar_expediente = int(
                                    clave_paciente_a_buscar_expediente
                                )
                                # rompe el bucle de la validacion de la clave del paciente
                                break
                            except ValueError:
                                print("Ingresa datos numericos")
                                continue
                        list_comp_claves = [
                            clave[0] for clave in resultado_pacientes_encontrados
                        ]

                        if clave_paciente_a_buscar_expediente in list_comp_claves:
                            cursor.execute(
                                "SELECT id_folio, fecha_cita, turno, hora_llegada, peso, estatura, presion_arterial, diagnostico, edad \
                                FROM citas  \
                                WHERE id_paciente = ? AND hora_llegada IS NOT NULL",
                                (clave_paciente_a_buscar_expediente,),
                            )

                            resultado_expediente = cursor.fetchall()

                            if resultado_expediente:
                                print(
                                    f"\n\n=======================================================EXPEDIENTE DEL PACIENTE {clave_paciente_a_buscar_expediente}======================================================"
                                )
                                print(
                                    f"{'FOLIO':^5}\t{'FECHA CITA':^5}\t{'TURNO':^5}\t{'HORA DE LLAGADA':<10}\t\t{'PESO':<10}\t{'ESTATURA':<10}\t{'PRESION ARTERIAL':^5}\t{'EDAD':<10}"
                                )
                                for tupla in resultado_expediente:
                                    (
                                        id_folio,
                                        fecha_cita,
                                        turno,
                                        hora_llegada,
                                        peso,
                                        estatura,
                                        presion_arterial,
                                        diagnostico,
                                        edad,
                                    ) = tupla
                                    print(
                                        f"{id_folio:^5}\t{fecha_cita:^5}\t{turno:^5}\t{hora_llegada}\t\t\t{peso}\t\t{estatura}\t\t{presion_arterial}\t\t\t{edad}"
                                    )
                                    print(f"Diagnostico: \n\t{diagnostico}")
                            else:
                                print(
                                    f"El paciente con la clave {clave_paciente_a_buscar_expediente} no tiene citas realizadas"
                                )
                        else:
                            print(
                                "El paciente con esta clave no se encuentra en el filtrado por nombres y apellidos"
                            )

                    else:
                        print("No hubo nunguna coincidencia")

                    # rompe el aviso
                    break

            else:
                print("No existen pacientes registrados")

            resultado = cursor.fetchall()
    except sqlite3.Error as e:
        print(e)
    finally:
        if conn:
            conn.close()


# =============


def estadisticos_demograficos():
    while True:
        print("\n\n==========ESTADISTICOS DESCRIPTIVOS==========")
        print(
            " 1. Por edad. \n 2. Por sexo. \n 3. Por edad y sexo. \n 4. Regresar al menu consultas y reportes."
        )

        while True:
            opcion_estadisticos_descriptivos = input("\nIngresa la opcion deseada: ")
            if opcion_estadisticos_descriptivos == "*":
                break
            if opcion_estadisticos_descriptivos == "":
                print("No puedes omitir este campo")
                continue
            try:
                opcion_estadisticos_descriptivos = int(opcion_estadisticos_descriptivos)
                # rompiendo el bucle para validar la opcion del submenu
                break
            except ValueError:
                print("Ingresa valores numericos")
                continue

        # 1. Por edad recibe un tango de edades
        if opcion_estadisticos_descriptivos == 1:
            estadisticos_demograficos_por_edad()
        elif opcion_estadisticos_descriptivos == 2:
            estadisticos_demograficos_por_sexo()
        elif opcion_estadisticos_descriptivos == 3:
            estadisticos_demograficos_por_edad_y_sexo()
        elif opcion_estadisticos_descriptivos == 4:
            break
        else:
            print("Selecciona una opcion entre 1-4")
            continue


# def listo
def estadisticos_demograficos_por_edad():
    # bucle del aviso
    while True:
        print("AVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")
        while True:
            edad_inicio = input("\nIngresa desde que edad deseas:")
            if edad_inicio == "*":
                break
            if edad_inicio == "":
                print("No puedes omitir este valor")
                continue

            try:
                edad_inicio = int(edad_inicio)
            except ValueError:
                print("Ingresa numeros enteros")
            if edad_inicio <= 0:
                print("La edad no puede tener valores negativos")
                continue

            break

        if edad_inicio == "*":
            break

        while True:
            edad_fin = input("Ingresa hasta que edad deseas: ")
            if edad_fin == "*":
                break
            if edad_fin == "":
                print("No puedes omitir este valor")
                continue

            try:
                edad_fin = int(edad_fin)
            except ValueError:
                print("Ingresa numeros enteros")
            if edad_fin <= 0:
                print("La edad no puede tener valores negativos")
                continue
            break

        if edad_fin == "*":
            break

        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT id_folio, peso, estatura \
                    FROM citas \
                    WHERE edad BETWEEN ? AND ?",
                    (edad_inicio, edad_fin),
                )

                resultado = cursor.fetchall()

        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()

        try:
            d_frame = pd.DataFrame(resultado, columns=["Folio", "Peso", "Estatura"])
            d_frame.set_index("Folio", inplace=True)
            if d_frame.empty:
                print("No se encontraron datos.")
            else:
                print(f"\n{d_frame}")
                conteo = d_frame.count()
                valor_minimo = d_frame.min()
                valor_maximo = d_frame.max()
                media = d_frame.mean()
                mediana = d_frame.median()
                desviacion_estandar = d_frame.std()

                # Mostrar resultados
                print("\nConteo: \n", conteo)
                print("\nValor mínimo:\n", valor_minimo)
                print("\nValor máximo:\n", valor_maximo)
                print(
                    "\nMedia aritmética:\n", media
                )  # suma de todos los numeros entre la cantidad de numeros "n"
                print("\nMediana:\n", mediana)
                print("\nDesviación estándar:\n", desviacion_estandar)

        except pd.errors.EmptyDataError:
            print("Los datos están vacíos, no se pudo crear el DataFrame.")

            # rompe el bucle del aviso
        break


# def listo
def estadisticos_demograficos_por_sexo():
    while True:
        print("AVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")
        while True:
            sexo = input("\nSeleccione algun sexo ( 'H'/ 'M' / 'N' ): ").upper()
            if sexo == "*":
                break
            if sexo == "":
                print("No puedes omitir este valor")
                continue
            if sexo in ["H", "M", "N"]:
                break
            else:
                print("Seleccione un sexo valido.")

        if sexo == "*":
            break

        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT c.id_folio, c.peso, c.estatura \
                    FROM citas c \
                    INNER JOIN registro_pacientes r ON r.id_paciente = c.id_paciente \
                    WHERE r.sexo = ?",
                    (sexo,),
                )

                resultado = cursor.fetchall()

                d_frame = pd.DataFrame(resultado, columns=["Folio", "Peso", "Estatura"])
                d_frame.set_index("Folio", inplace=True)
                if d_frame.empty:
                    print("No se encontraron datos con este sexo.")
                else:
                    print(f"\n{d_frame}")
                    conteo = d_frame.count()
                    valor_minimo = d_frame.min()
                    valor_maximo = d_frame.max()
                    media = d_frame.mean()
                    mediana = d_frame.median()
                    desviacion_estandar = d_frame.std()

                    # Mostrar resultados
                    print("\nConteo: \n", conteo)
                    print("\nValor mínimo:\n", valor_minimo)
                    print("\nValor máximo:\n", valor_maximo)
                    print(
                        "\nMedia aritmética:\n", media
                    )  # suma de todos los numeros entre la cantidad de numeros "n"
                    print("\nMediana:\n", mediana)
                    print("\nDesviación estándar:\n", desviacion_estandar)

        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()

        break


# def pendiente
def estadisticos_demograficos_por_edad_y_sexo():
    while True:
        print("AVISO: \n\tPara cancelar el proceso en cualquier opcion ingresa '*'")
        while True:
            edad_inicio = input("\nIngresa desde que edad deseas:")
            if edad_inicio == "*":
                break
            if edad_inicio == "":
                print("No puedes omitir este valor")
                continue
            try:
                edad_inicio = int(edad_inicio)
            except ValueError:
                print("Ingresa numeros enteros")
            if edad_inicio <= 0:
                print("La edad no puede tener valores negativos")
                continue
            break
        if edad_inicio == "*":
            break

        while True:
            edad_fin = input("Ingresa hasta que edad deseas: ")
            if edad_fin == "*":
                break
            if edad_fin == "":
                print("No puedes omitir este valor")
                continue
            try:
                edad_fin = int(edad_fin)
            except ValueError:
                print("Ingresa numeros enteros")
            if edad_fin <= 0:
                print("La edad no puede tener valores negativos")
                continue
            break
        if edad_fin == "*":
            break

        while True:
            sexo = input("\nSeleccione algun sexo ( 'H'/ 'M' / 'N' ): ").upper()
            if sexo == "*":
                break
            if sexo == "":
                print("No puedes omitir este valor")
                continue
            if sexo in ["H", "M", "N"]:
                pass
            else:
                print("Seleccione un sexo valido.")

            break
        if sexo == "*":
            break

        try:
            with sqlite3.connect("base_datos_ev3.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT c.id_folio, c.peso, c.estatura \
                    FROM citas c \
                    INNER JOIN registro_pacientes r ON r.id_paciente = c.id_paciente \
                    WHERE c.edad BETWEEN ? AND ? AND r.sexo = ?",
                    (edad_inicio, edad_fin, sexo),
                )

                resultado = cursor.fetchall()

                d_frame = pd.DataFrame(resultado, columns=["Folio", "Peso", "Estatura"])
                d_frame.set_index("Folio", inplace=True)
                if d_frame.empty:
                    print("No se encontraron datos.")
                else:
                    print(f"\n{d_frame}")
                    conteo = d_frame.count()
                    valor_minimo = d_frame.min()
                    valor_maximo = d_frame.max()
                    media = d_frame.mean()
                    mediana = d_frame.median()
                    desviacion_estandar = d_frame.std()

                    # Mostrar resultados
                    print("\nConteo: ", conteo)
                    print("\nValor mínimo:\n", valor_minimo)
                    print("\nValor máximo:\n", valor_maximo)
                    print(
                        "\nMedia aritmética:\n", media
                    )  # suma de todos los numeros entre la cantidad de numeros "n"
                    print("\nMediana:\n", mediana)
                    print("\nDesviación estándar:\n", desviacion_estandar)

        except sqlite3.Error as e:
            print(e)
        finally:
            if conn:
                conn.close()
        break


limpiar_consola()
# estadisticos_demograficos_por_edad_y_sexo()

# estadisticos_demograficos_por_sexo()

# estadisticos_demograficos_por_edad()

# busqueda_por_clave_paciente()

# busqueda_por_apellidos_y_nombres()

# busqueda_por_clave_paciente()

# listado_completo_pacientes()

# reportes_de_citas()

# reporte_de_citas_por_paciente()

# reporte_de_citas_por_periodo()

# bloque_principal()

# creacion_tablas()

# registrar_pacientes()  # Funcionando

# programar_citas()

# realizar_cita()

# cancelar_cita_busqueda_por_fecha()

# cancelar_cita_busqueda_por_paciente()

# sub_menu_cancelacion_citas()

# sub_menu_consultas_y_reportes()

# reportes_de_citas()

bloque_principal()
